from openpyxl import  Workbook, load_workbook
from openpyxl.styles import Font, colors, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import csv
import os
import yaml
subclass = ['vision','language']
def col(startcol, offset):
    return chr(ord(startcol)+offset)
def row(startrow, offset):
    return str(startrow+offset)

sr = 2 #start row
sc = 'A' #start column
leftalign = Alignment(horizontal='left',vertical='center')

def init_table(target, tablename):
    wb = Workbook()
    ws = wb.active
    ws.title = target
    ws[col(sc,0)+row(sr,0)] = 'NetClass'
    ws.merge_cells(col(sc,0)+row(sr,0)+':'+col(sc,0)+row(sr,1))
    ws[col(sc,0)+row(sr,0)].alignment = Alignment(horizontal='center', vertical='center')

    ws[col(sc,1)+row(sr,0)] = 'NetName'
    ws.merge_cells(col(sc,1)+row(sr,0)+':'+col(sc,1)+row(sr,1))
    ws[col(sc,1)+row(sr,0)].alignment = Alignment(horizontal='center', vertical='center')

    ws[col(sc,2)+row(sr,0)] = 'Shape'
    ws.merge_cells(col(sc,2)+row(sr,0)+':'+col(sc,2)+row(sr,1))
    ws[col(sc,2)+row(sr,0)].alignment = Alignment(horizontal='center', vertical='center')

    if target=='BM1684':
        ws.title = 'BM1684'
        ws[col(sc,3)+row(sr,0)] = 'BM1684 Benchmark(qps)'
        #fp32,int8 1b~16b
        ws.merge_cells(col(sc,3)+row(sr,0)+':'+col(sc,7)+row(sr,0))
        ws[col(sc,3)+row(sr,0)].alignment = Alignment(horizontal='center', vertical='center')
        ws[col(sc,3)+row(sr,1)] = 'fp32'
        ws[col(sc,4)+row(sr,1)] = 'int8 1batch'
        ws[col(sc,5)+row(sr,1)] = 'int8 4batch'
        ws[col(sc,6)+row(sr,1)] = 'int8 8batch'
        ws[col(sc,7)+row(sr,1)] = 'int8 16batch'

        ws[col(sc,8)+row(sr,0)] = 'Gops'
        ws.merge_cells(col(sc,8)+row(sr,0)+':'+col(sc,8)+row(sr,1))
        ws[col(sc,8)+row(sr,0)].alignment = Alignment(horizontal='center', vertical='center')

        ws[col(sc,9)+row(sr,0)] = 'Resource'
        ws.merge_cells(col(sc,9)+row(sr,0)+':'+col(sc,9)+row(sr,1))
        ws[col(sc,9)+row(sr,0)].alignment = Alignment(horizontal='center', vertical='center')


    else: #BM1684X
        ws.title = 'BM1684X'
        ws[col(sc,3)+row(sr,0)] = 'BM1684X@1Ghz Benchmark(qps)'
        #fp32, fp16, int8 1b~16b
        ws.merge_cells(col(sc,3)+row(sr,0)+':'+col(sc,8)+row(sr,0))
        ws[col(sc,3)+row(sr,0)].alignment = Alignment(horizontal='center', vertical='center')
        ws[col(sc,3)+row(sr,1)] = 'fp32'
        ws[col(sc,4)+row(sr,1)] = 'fp16'
        ws[col(sc,5)+row(sr,1)] = 'int8 1batch'
        ws[col(sc,6)+row(sr,1)] = 'int8 4batch'
        ws[col(sc,7)+row(sr,1)] = 'int8 8batch'
        ws[col(sc,8)+row(sr,1)] = 'int8 16batch'

        ws[col(sc,9)+row(sr,0)] = 'Gops'
        ws.merge_cells(col(sc,9)+row(sr,0)+':'+col(sc,9)+row(sr,1))
        ws[col(sc,9)+row(sr,0)].alignment = Alignment(horizontal='center', vertical='center')

        ws[col(sc,10)+row(sr,0)] = 'Resource'
        ws.merge_cells(col(sc,10)+row(sr,0)+':'+col(sc,10)+row(sr,1))
        ws[col(sc,10)+row(sr,0)].alignment = Alignment(horizontal='center', vertical='center')

    wb.save(tablename)

def adjust_sheet(filename):
    wb=load_workbook(filename)
    for sheet in wb.sheetnames:
      ws=wb[sheet]
      df=pd.read_excel(filename,sheet).fillna('-')
      df.loc[len(df)]=list(df.columns)
      for column in df.columns:
        index=list(df.columns).index(column)
        letter=get_column_letter(index+1)
        collen=df[column].apply(lambda x:len(str(x).encode())).max()
        ws.column_dimensions[letter].width=collen+2

    #left side align
    if ws.title == 'BM1684X':
        endsc = 8
    else:
        endsc = 7

    colnum = 3
    #print(ws.max_row)
    for column in ws[col(sc,3)+':'+ col(sc,endsc)]:
        for rownum in range(ws.max_row):
            #print(colnum, rownum)
            rownum += 1
            if (rownum > (sr+1)):
                #print(colnum, rownum, ws.cell(rownum, colnum).value)
                ws[col(sc,colnum)+str(rownum)].alignment = Alignment(horizontal='left', vertical='center')
        colnum+=1

    #bind the cells with same value
    rowcnt = 1
    preval = ''
    postval = ''
    startrow = []
    endrow = []
    first = True
    for rows in ws[col(sc, 0)]:
        if (rows.value in subclass) and first:
            startrow.append(rowcnt)
            first = False
            preval = rows.value
            postval = ws.cell(rowcnt+1, (ord(sc) - ord('A'))+1).value
            #print(preval, postval)
        postval = ws.cell(rowcnt+1, (ord(sc) - ord('A'))+1).value
        if (preval!=postval) and (preval != ''):
            endrow.append(rowcnt)
            first = True
        rowcnt += 1
    #print(startrow,endrow)
    for i in range(len(startrow)):
        ws.merge_cells(col(sc,0)+str(startrow[i])+':'+col(sc,0)+str(endrow[i]))
        ws[col(sc,0)+str(startrow[i])].alignment = Alignment(horizontal='left', vertical='center')

    wb.save(filename)

def throughput(time, batchsize):
    fps = 1000/(float(time)/batchsize)
    #print(time, batchsize, fps)
    return float('%.2f'%fps)

def find_class(netname, classes):
    #print(classes)
    if classes is not None:
      for bind in classes:
        if bind[1] == netname:
          return bind[0]

def analyze_stat(statpath, class_type):
    bench = []
    item = dict()
    with open(statpath, 'r') as file:
      csv_file = csv.DictReader(file)
      pre_netname = ''
      new_netname = ''
      for row in csv_file:
        new_netname = row['name']
        #print(pre_netname, new_netname)
        if pre_netname != new_netname:
          tmp = item.copy()
          if tmp!= {}:
            bench.append(tmp)
          item.clear()
          item={'class':'','name':'','shape':'','fp32':'N/A','fp16':'N/A','int8-1b':'N/A', \
                'int8-4b':'N/A','int8-8b':'N/A','int8-16b':'N/A','gops':'N/A'}
          shape = row['shape'].split('x')[1:]
          dims = '*'.join(shape)
          item['class'] = find_class(row['name'], class_type)
          item['name'] = row['name']
          item['shape'] = dims
          item['gops'] = row['gops']
          pre_netname = new_netname
          time = 'time(ms)'
          if row['prec']=='FP32':
            item['fp32'] = throughput(row[time], 1)
          elif ((row['prec'] == 'FP16') or (row['prec'] == 'BF16')):
            item['fp16'] = throughput(row[time], 1)
          else:
            if(row['shape'].split('x')[0]=='1'):
              item['int8-1b'] = throughput(row[time], 1)
            elif(row['shape'].split('x')[0]=='4'):
              item['int8-4b'] = throughput(row[time], 4)
            elif(row['shape'].split('x')[0]=='8'):
              item['int8-8b'] = throughput(row[time], 8)
            elif(row['shape'].split('x')[0]=='16'):
              item['int8-16b'] = throughput(row[time], 16)
        else:
          if row['prec']=='FP32':
            item['fp32'] = throughput(row[time], 1)
          elif ((row['prec'] == 'FP16') or (row['prec'] == 'BF16')):
            item['fp16'] = throughput(row[time], 1)
          else:
            if(row['shape'].split('x')[0]=='1'):
              item['int8-1b'] = throughput(row[time], 1)
            elif(row['shape'].split('x')[0]=='4'):
              item['int8-4b'] = throughput(row[time], 4)
            elif(row['shape'].split('x')[0]=='8'):
              item['int8-8b'] = throughput(row[time], 8)
            elif(row['shape'].split('x')[0]=='16'):
              item['int8-16b'] = throughput(row[time], 16)

      tmp = item.copy()
      bench.append(tmp)
      #print(bench)
      return bench

def fill_table(bench, tablename, target):
    wb = load_workbook(tablename)
    ws = wb.active
    #ws.append(['name', 'shape', 'fp32', 'fp16','int8-1batch','int8-4batch','int8-8batch','int8-16batch'])
    for item in bench:
        #print(item)
        if target=='BM1684X':
            ws.append([item['class'],item['name'], item['shape'], item['fp32'],item['fp16'], \
                     item['int8-1b'],item['int8-4b'],item['int8-8b'], \
                     item['int8-16b'], item['gops']])
        else:
            ws.append([item['class'],item['name'], item['shape'], item['fp32'], \
                     item['int8-1b'],item['int8-4b'],item['int8-8b'], \
                     item['int8-16b'], item['gops']])
    wb.save(tablename)

def read_config(path):
    '''
    fn = os.path.join(path, 'config.yaml')
    if not os.path.exists(fn):
        print(f'No config in {path}')
        return
    '''
    with open(path) as f:
        return yaml.load(f, yaml.Loader)

def get_class(zoo_path):
    results = []
    all=os.walk(zoo_path)
    #print(zoo_path)
    for p, ds, fs in all:
        #print(p)
        for f in fs:
            fullname = os.path.join(p,f)
            if fullname.endswith('config.yaml'):
                if zoo_path.endswith('/'):
                    subpath = fullname.replace(zoo_path, '')
                else:
                    subpath = fullname.replace(zoo_path+'/','')
                folders = subpath.split('/')
                #print(folders)
                config = read_config(fullname)
                item = dict()
                if 'name' in config:
                    #print(config['name'],config['gops'])
                    item['name'] = config['name']
                    #print(folders[0])
                    if folders[0] in subclass:
                        item['class'] = folders[0]
                    #item['gops'] = config['gops']
                    results.append((item['class'],item['name']))
    return results

def main():
    import argparse
    parser = argparse.ArgumentParser(description='model-zoo benchmark tool')
    parser.add_argument('--stat', type=str, default='', required=True,\
                        help='the path/to/stat.csv')
    parser.add_argument('--target', type=str, default='BM1684X', \
                        help='the device type, BM1684 or BM1684X')
    parser.add_argument('--model_zoo', type=str, default='', required=True, \
                        help='the path/to/model-zoo')
    parser.add_argument('--table_name', type=str, default='formatted_result.xlsx', \
                        help='the output file name')
    args = parser.parse_args()
    classtype = get_class(args.model_zoo)
    init_table(args.target, args.table_name)
    output = analyze_stat(args.stat, classtype)
    fill_table(output, args.table_name, args.target)
    adjust_sheet(args.table_name)


if __name__ == '__main__':
    main()
